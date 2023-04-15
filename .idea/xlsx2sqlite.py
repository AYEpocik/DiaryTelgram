# Импортируем необходимые библиотеки
import openpyxl
import sqlite3
import os

table_path = '../Database/school.xlsx'

# Открываем файл .xlsx с помощью openpyxl
wb = openpyxl.load_workbook(table_path)

# Создаем имя файла .db с тем же именем, что и .xlsx, но с другим расширением
db_name = os.path.splitext(table_path)[0] + ".db"

# Создаем соединение с файлом .db с помощью sqlite3
conn = sqlite3.connect(db_name)

# Создаем курсор для выполнения SQL-запросов
cur = conn.cursor()

# Перебираем все имена листов в файле .xlsx
for sheet_name in wb.sheetnames:
    # Получаем объект листа по имени
    sheet = wb[sheet_name]

    # Получаем количество строк и столбцов в листе
    rows = sheet.max_row
    cols = sheet.max_column

    # Создаем список для хранения имен полей
    fields = []

    # Читаем имена полей из первой строки листа и добавляем их в список
    for col in range(1, cols + 1):
        field = sheet.cell(row=1, column=col).value
        fields.append(field)

    # Создаем SQL-запрос для создания таблицы с тем же именем, что и лист, и теми же полями
    sql_create = f"CREATE TABLE IF NOT EXISTS {sheet_name} ("

    # Добавляем имена полей и типы данных в SQL-запрос
    for field in fields:
        sql_create += f"{field} TEXT,"

    # Удаляем последнюю запятую из SQL-запроса
    sql_create = sql_create[:-1]

    # Добавляем закрывающую скобку в SQL-запрос
    sql_create += ")"

    # Выполняем SQL-запрос для создания таблицы
    cur.execute(sql_create)

    # Создаем SQL-запрос для вставки данных из листа в таблицу
    sql_insert = f"INSERT INTO {sheet_name} VALUES ("

    # Добавляем знаки вопроса в SQL-запрос для каждого поля
    for field in fields:
        sql_insert += "?,"

    # Удаляем последнюю запятую из SQL-запроса
    sql_insert = sql_insert[:-1]

    # Добавляем закрывающую скобку в SQL-запрос
    sql_insert += ")"

    # Читаем данные из листа, начиная со второй строки, и добавляем их в список кортежей
    data = []
    for row in range(2, rows + 1):
        record = []
        for col in range(1, cols + 1):
            value = sheet.cell(row=row, column=col).value
            record.append(value)
        data.append(tuple(record))

    # Выполняем SQL-запрос для вставки данных в таблицу с помощью метода executemany
    cur.executemany(sql_insert, data)

# Сохраняем изменения в файле .db с помощью метода commit
conn.commit()

# Закрываем соединение с файлом .db с помощью метода close
conn.close()

#Выводим сообщение об успешном завершении программы
print(f"Файл {db_name} успешно создан.")
